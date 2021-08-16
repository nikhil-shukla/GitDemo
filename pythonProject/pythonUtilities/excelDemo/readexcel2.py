import openpyxl

wb=openpyxl.load_workbook("Data.xlsx")

sh1=wb['Names']
row=sh1.max_row
column=sh1.max_column

for i in range(1,row+1): #+1 for all values
    for j in range(1,column+1):
        print(sh1.cell(i,j).value)

sh1.cell(row=5,column=1,value='pytest')
sh1.cell(row=5,column=2,value='uk')
sh1.cell(row=5,column=3,value=49)

wb.save('Report.xlsx') #new file so that data will be appended not overrided
