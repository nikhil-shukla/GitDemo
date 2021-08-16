from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb=Workbook()
print(wb.active.title)
print(wb.sheetnames)
wb['Sheet'].title="Report of automation"
sh1=wb.active
sh1['A1'].value="Name"
sh1['B1'].value="Status"
sh1['A2'].value="Python"
sh1['B2'].value="Active"
sh1['B2'].fill=PatternFill("solid",fgColor="00FF00")
sh1['A3'].value="Java"
sh1['B3'].value="Inactive"
sh1['B3'].fill=PatternFill("solid",fgColor="FF0000")
wb.save("FinalReport.xlsx")