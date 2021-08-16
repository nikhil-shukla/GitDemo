import openpyxl

wb=openpyxl.load_workbook("Data.xlsx")
#print(type(wb))
sheets=wb.sheetnames
print(sheets)
#print(wb.active.title)

sh1=wb['Names']
print(type(sh1))

data=sh1['B2'].value
print(data)


#option1
print(wb['Names']['A2'].value)


#option2
print(sh1.cell(3,2).value)
print(sh1.cell(3,3).value)

sh2=wb['Marks']
print(sh2.cell(2,1).value)
print(sh2.cell(3,2).value)


#option3
c=sh2.cell(row=2,column=2)
print(type(c))
print(sh2.cell(row=2,column=2).value)

#deprecated
print(wb.get_sheet_by_name('Names').cell(row=4,column=1).value)



